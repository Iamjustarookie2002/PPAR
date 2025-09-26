import pandas as pd
import re
import colorsys
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU

def process_excel_report(df, excel_filename, visits_df, manual_patient_data=None):
    """
    Main function that creates the Excel file with both sheets.
    This is the ONLY function accessible to main in app.py.
    
    Args:
        df: DataFrame with the data from FILES_DAT sheet
        excel_filename: Output Excel filename
        visits_df: DataFrame with patient data from VISITS sheet
        manual_patient_data: Dictionary with manual patient data (optional)
    """
    try:
        wb = Workbook()
        
        # Create Sheet2 first and process it with all data
        ws2 = wb.create_sheet("Sheet2")
        num_data_rows = process_sheet2_data(df, ws2)
        
        # Calculate the row numbers for summary tables in Sheet2
        summary_start_row = num_data_rows + 6  # Main data + gap + summary table start
        forelimb_start_row = num_data_rows + 10  # SI values are always at rows 16 and 17 in Sheet2
        
        # Create Sheet1 and process it with formulas referencing Sheet2
        ws1 = wb.active
        ws1.title = "Sheet1"
        process_sheet1_data(ws1, visits_df, summary_start_row, forelimb_start_row, manual_patient_data)
        
        # Save the workbook
        wb.save(excel_filename)
        
    except Exception as e:
        print(f"Error in process_excel_report: {e}")
        # Create a minimal workbook with error message
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error"
            ws.cell(row=1, column=1, value=f"Error processing file: {str(e)}")
            wb.save(excel_filename)
        except:
            # If even saving fails, create a simple text file
            with open(excel_filename.replace('.xlsx', '_error.txt'), 'w') as f:
                f.write(f"Error processing file: {str(e)}")
        raise e

def process_sheet1_data(ws1, visits_df, summary_start_row, forelimb_start_row, manual_patient_data=None):
    """
    Process Sheet1 - populate patient data from VISITS sheet and manual inputs, add summary averages table from Sheet2.
    
    Args:
        ws1: Worksheet object for Sheet1
        visits_df: DataFrame with patient data from VISITS sheet
        manual_patient_data: Dictionary with manual patient data (optional)
    """
    # Set up the dashboard layout
    ws1.row_dimensions[1].height = 30  # Set title row height
    
    # Add main title
    ws1.merge_cells(f'B1:C1')
    title_cell = ws1.cell(row=1, column=2, value="PVH gait lab report")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add current date beside the title (small format)
    current_date = datetime.now().strftime("%d/%m/%Y")
    date_cell = ws1.cell(row=1, column=4, value=current_date)
    date_cell.font = Font(size=10)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    patient_info_row = 3
    
    # Extract patient data from VISITS sheet (first row, excluding N3, N2, N1 columns)
    if not visits_df.empty:
        # Get the first row of data (excluding N3, N2, N1 columns)
        patient_data = visits_df.iloc[0]
        
        # Extract patient information from VISITS sheet
        first_name = patient_data.get('First name', '')
        last_name = patient_data.get('Last name', '')
        gender = patient_data.get('Gender', '')
        visits_id = patient_data.get('ID', '')
        date_of_birth = patient_data.get('Date of birth', '')
        date_of_visit = patient_data.get('Date of visit', '')
        body_mass = patient_data.get('Body mass [kg]', '')
        
        # Create full name
        full_name = f"{first_name} {last_name}".strip()
        
        # Format date of visit
        if date_of_visit:
            if isinstance(date_of_visit, str):
                visit_date = date_of_visit
            else:
                # Handle datetime objects - remove time component
                visit_date = date_of_visit.strftime("%m/%d/%Y") if hasattr(date_of_visit, 'strftime') else str(date_of_visit)
        else:
            visit_date = ''
        
        # Format body weight
        body_weight = f"{body_mass} kg" if body_mass else ''
    else:
        # Default values if no data
        first_name = ''
        last_name = ''
        gender = ''
        visits_id = ''
        date_of_birth = ''
        date_of_visit = ''
        body_mass = ''
        full_name = ''
        visit_date = ''
        body_weight = ''
    
    signalment = []
    # Override with manual patient data if provided
    if manual_patient_data:
        # Use manual data if provided, otherwise use VISITS sheet data
        species = manual_patient_data.get('species', '').strip()
        breed = manual_patient_data.get('breed', '').strip()
        color = manual_patient_data.get('color', '').strip()
        manual_id = manual_patient_data.get('purdue_id', '').strip()
        primary_dvm = manual_patient_data.get('primary_dvm', '').strip()
        
        # Build enhanced signalment with manual data
        signalment_parts = []
        if species:
            signalment_parts.append(f"Species: {species}")
        if breed:
            signalment_parts.append(f"Breed: {breed}")
        if color:
            signalment_parts.append(f"Color: {color}")
        
        signalment = signalment_parts  # Keep as list

    # Add gender, DOB, and age to signalment (regardless of manual data)
    if gender:
        signalment.append(f"Gender: {gender}")
    if date_of_birth:
        # Convert Timestamp to string if needed
        if hasattr(date_of_birth, 'strftime'):
            dob_str = date_of_birth.strftime("%m/%d/%Y")
        else:
            dob_str = str(date_of_birth)
        signalment.append(f"DOB: {dob_str}")  # Use dob_str, not date_of_birth
        
        # Also convert date_of_visit if it exists
        visit_str = None
        if date_of_visit:
            if hasattr(date_of_visit, 'strftime'):
                visit_str = date_of_visit.strftime("%m/%d/%Y")
            else:
                visit_str = str(date_of_visit)
        
        years, months = calculate_age_years_months(dob_str, visit_str)
        if years is not None and months is not None:
            signalment.append(f"Age: {years}y {months}m")

    # Convert signalment list to string
    signalment = ", ".join(signalment) if signalment else ""
    
    # Add patient information labels in column 1 and values in column 3
    patient_labels = ["Name:", "Signalment:", "MR-ID:", "VisitDate:", "BW:", "PrimaryDVM:", "Purdue-ID:"]
    patient_values = [full_name, signalment, visits_id, visit_date, body_weight, primary_dvm, manual_id]
    ws1.merge_cells(f'B{patient_info_row+1}:H{patient_info_row+1}')
    
    for row_idx, (label, value) in enumerate(zip(patient_labels, patient_values), patient_info_row):
        # Add label in column 1
        label_cell = ws1.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True, size=12)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Add value in column 2
        value_cell = ws1.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(size=12)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # No image insertion - leave the area empty or add a placeholder
    
    # Insert the fixed DogTopView.png above the summary table
    try:
        # Try to insert DogTopView.png
        dog_img = Image('DogTopView.png')
        
        # Size the image to fit above the summary table
        dog_img.width = 130  # 1.5 columns wide
        dog_img.height = 400  # Maintain aspect ratio
        dog_image_row = patient_info_row + 9
        # Position above the summary table
        ws1.add_image(dog_img, f'B{dog_image_row}')
        
    except Exception as e:
        # If DogTopView.png is not found, add a placeholder
        placeholder_cell = ws1.cell(row=dog_image_row, column=2, value="[DogTopView.png not found]")

    cell_lf = ws1.cell(row=dog_image_row+5, column=1)
    cell_lf.fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
    cell_lf.value = f"=Sheet2!E{summary_start_row + 1}"
    
    
    # C13 - Rt. Forelimb color (light red)
    cell_rf = ws1.cell(row=dog_image_row+5, column=3)
    cell_rf.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    cell_rf.value = f"=Sheet2!E{summary_start_row + 3}"
    
    # A23 - Lt. Hindlimb color (light green)
    cell_lh = ws1.cell(row=dog_image_row+13, column=1)
    cell_lh.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    cell_lh.value = f"=Sheet2!E{summary_start_row + 2}"
    
    # B23 - Rt. Hindlimb color (light orange)
    cell_rh = ws1.cell(row=dog_image_row+13, column=3)
    cell_rh.fill = PatternFill(start_color='FFD699', end_color='FFD699', fill_type='solid')
    cell_rh.value = f"=Sheet2!E{summary_start_row + 4}"
    # Add summary averages table from Sheet2 below the DogTopView image
    # Start from row 14 to give space for the DogTopView image above
    start_row = dog_image_row + 20
        
    # Create summary table headers
    summary_headers = ["", "%BW", "VI [%BW*s]", "Contact time [ms]", "Weight bearing"]
    for col_idx, header in enumerate(summary_headers):
        header_cell = ws1.cell(row=start_row + 1, column=col_idx + 1, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Add grey fill to the metric headers (skip the first empty column)
        if col_idx > 0:  # Skip the first empty column
            header_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Add summary data with formulas referencing Sheet2 summary table
    # Use the passed row numbers for simple cell references
    
    # LF Summary - use simple cell references
    ws1.cell(row=start_row + 2, column=1, value="Lt. Forelimb").font = Font(bold=True)
    # Add blue fill to Lt. Forelimb
    lt_forelimb_cell = ws1.cell(row=start_row + 2, column=1)
    lt_forelimb_cell.fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
    ws1.cell(row=start_row + 2, column=2, value=f"=Sheet2!B{summary_start_row + 1}")
    ws1.cell(row=start_row + 2, column=3, value=f"=Sheet2!C{summary_start_row + 1}")
    ws1.cell(row=start_row + 2, column=4, value=f"=Sheet2!D{summary_start_row + 1}")
    ws1.cell(row=start_row + 2, column=5, value=f"=Sheet2!E{summary_start_row + 1}")
    
    # RF Summary - use simple cell references
    ws1.cell(row=start_row + 3, column=1, value="Rt. Forelimb").font = Font(bold=True)
    # Add red fill to Rt. Forelimb
    rt_forelimb_cell = ws1.cell(row=start_row + 3, column=1)
    rt_forelimb_cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    ws1.cell(row=start_row + 3, column=2, value=f"=Sheet2!B{summary_start_row + 3}")
    ws1.cell(row=start_row + 3, column=3, value=f"=Sheet2!C{summary_start_row + 3}")
    ws1.cell(row=start_row + 3, column=4, value=f"=Sheet2!D{summary_start_row + 3}")
    ws1.cell(row=start_row + 3, column=5, value=f"=Sheet2!E{summary_start_row + 3}")

    # LH Summary - use simple cell references
    ws1.cell(row=start_row + 4, column=1, value="Lt. Hindlimb").font = Font(bold=True)
    # Add green fill to Lt. Hindlimb
    lt_hindlimb_cell = ws1.cell(row=start_row + 4, column=1)
    lt_hindlimb_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    ws1.cell(row=start_row + 4, column=2, value=f"=Sheet2!B{summary_start_row + 2}")
    ws1.cell(row=start_row + 4, column=3, value=f"=Sheet2!C{summary_start_row + 2}")
    ws1.cell(row=start_row + 4, column=4, value=f"=Sheet2!D{summary_start_row + 2}")
    ws1.cell(row=start_row + 4, column=5, value=f"=Sheet2!E{summary_start_row + 2}")
    
    
    # RH Summary - use simple cell references
    ws1.cell(row=start_row + 5, column=1, value="Rt. Hindlimb").font = Font(bold=True)
    # Add orange fill to Rt. Hindlimb
    rt_hindlimb_cell = ws1.cell(row=start_row + 5, column=1)
    rt_hindlimb_cell.fill = PatternFill(start_color='FFD699', end_color='FFD699', fill_type='solid')
    ws1.cell(row=start_row + 5, column=2, value=f"=Sheet2!B{summary_start_row + 4}")
    ws1.cell(row=start_row + 5, column=3, value=f"=Sheet2!C{summary_start_row + 4}")
    ws1.cell(row=start_row + 5, column=4, value=f"=Sheet2!D{summary_start_row + 4}")
    ws1.cell(row=start_row + 5, column=5, value=f"=Sheet2!E{summary_start_row + 4}")
    
    abbreviations_row = start_row+6
    ws1.merge_cells(f'D{abbreviations_row}:E{abbreviations_row}')
    abbreviations_cell = ws1.cell(row=abbreviations_row, column=4, value="*BW: body weight, VI: vertical impulse")
    abbreviations_cell.font = Font(italic=True, size=10, underline='single')
    abbreviations_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Color the 4 specified cells with the same colors as limb labels
    # A13 - Lt. Forelimb color (light blue)
    

    # Add Forelimb/Hindlimb summary below the main summary table
    forelimb_start_row_sheet1 = dog_image_row + 7
    ws1.merge_cells(f'C{forelimb_start_row_sheet1}:D{forelimb_start_row_sheet1}')
    forelimb_cell = ws1.cell(row=forelimb_start_row_sheet1, column=3, value="Symmetry Index (SI)")
    forelimb_cell.font = Font(bold=True, size=14)
    forelimb_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Add forelimb/hindlimb data with original formulae
    forelimb_label = ws1.cell(row=forelimb_start_row_sheet1 + 1, column=3, value="Forelimb")
    forelimb_label.font = Font(bold=True)
    forelimb_label.alignment = Alignment(horizontal="right", vertical="center")
    ws1.cell(row=forelimb_start_row_sheet1 + 1, column=4, value=f"=Sheet2!B{forelimb_start_row + 2}")
    
    hindlimb_label = ws1.cell(row=forelimb_start_row_sheet1 + 2, column=3, value="Hindlimb")
    hindlimb_label.font = Font(bold=True)
    hindlimb_label.alignment = Alignment(horizontal="right", vertical="center")
    ws1.cell(row=forelimb_start_row_sheet1 + 2, column=4, value=f"=Sheet2!B{forelimb_start_row + 3}")
    
    abbreviations2_row = forelimb_start_row_sheet1+3
    ws1.merge_cells(f'C{abbreviations2_row}:D{abbreviations2_row}')
    abbreviations_cell2 = ws1.cell(row=abbreviations2_row, column=3, value="*lower SI means more symmetric")
    abbreviations_cell2.font = Font(italic=True, size=10, underline='single')
    abbreviations_cell2.alignment = Alignment(horizontal="right", vertical="center")
    # Set column widths dynamically based on header and content lengths
    # Get the maximum width needed for each column
    column_widths = {}
    # Apply -10 adjustment to all columns
    column_widths['A'] = 11  # 15
    column_widths['B'] = 16  # 10
    summary_headers = ["", "Maximum force [%BW]", "Force-time integral [%BW*s]", "Contact time/TO [ms]", "Weight bearing [%]"]
    for col_idx, header in enumerate(summary_headers):
        col_letter = get_column_letter(col_idx + 1)
        if col_letter not in column_widths:
            header_length = len(str(header))
            # Subtract 10 from the calculated width
            column_widths[col_letter] = max(header_length - 10, 15)  # Minimum width of 8
    for col_letter, width in column_widths.items():
        ws1.column_dimensions[col_letter].width = width

def process_sheet2_data(df, ws2):
    """
    Process and format Sheet2 with data processing, coloring, and additional columns.
    """
    processed_df = process_original_excel_data(df)
    num_data_rows = len(processed_df)

    # Write DataFrame to Sheet2 with proper formatting
    # Write headers first with bold formatting
    for col_idx, col_name in enumerate(processed_df.columns):
        header_cell = ws2.cell(row=1, column=col_idx + 1, value=col_name)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows with center alignment
    for row_idx, (_, row_data) in enumerate(processed_df.iterrows(), 2): # Start from row 2 for data
        for col_idx, value in enumerate(row_data):
            data_cell = ws2.cell(row=row_idx, column=col_idx + 1, value=value)
            data_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add additional columns to the right
    add_additional_columns_to_sheet2(ws2, num_data_rows)
    
    # Calculate and populate weight bearing percentages
    write_weight_bearing_formulae(ws2, num_data_rows)
    
    # Write asymmetry index formulae
    write_asymmetry_formulae(ws2, num_data_rows)
    
    # Apply coloring to Data Source and Weight bearing columns
    apply_coloring(ws2, num_data_rows)
    
    # Add summary table with averages
    add_summary_averages_table(ws2, num_data_rows)
    
    # Add forelimb/hindlimb asymmetry summary table
    add_forelimb_hindlimb_summary(ws2, num_data_rows)
    
    # Set column widths based on content
    set_column_widths(ws2, processed_df)

    return num_data_rows

def process_original_excel_data(df):
    """Process and filter the original Excel data with robust error handling."""
    try:
        processed_df = df.copy()

        # Filter out rows with '.dat' in "File short name"
        if "File short name" in processed_df.columns:
            processed_df = processed_df[~processed_df["File short name"].str.endswith(".dat", na=False)]

        # Select and rename required columns with error handling
        column_mapping = {
            "File comment": "Data Source",
            "Maximum force (normalized to BW) /Total object/ [%BW]": "Maximum force [%BW]",
            "Force-time integral (normalized to BW) /Total object/ [%BW*s]": "Force-time integral [%BW*s]",
            "Contact time/TO [ms]": "Contact time/TO [ms]",
        }
        
        # Check which columns exist and only use available ones
        available_columns = []
        for original_col, new_col in column_mapping.items():
            if original_col in processed_df.columns:
                available_columns.append(original_col)
            else:
                print(f"Warning: Column '{original_col}' not found in data")
        
        if not available_columns:
            raise ValueError("No required columns found in the Excel file")
        
        processed_df = processed_df[available_columns].rename(columns=column_mapping)

        # Handle missing data by filling with appropriate defaults
        for col in processed_df.columns:
            if col != "Data Source":  # Don't fill text columns
                # Fill numeric columns with 0 for missing values
                processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0)

        # Convert LF1 -> LF_1, LH2 -> LH_2, etc. with error handling
        if "Data Source" in processed_df.columns:
            processed_df["Data Source"] = processed_df["Data Source"].apply(
                lambda x: re.sub(r'([A-Z]+)(\d+)', r'\1_\2', str(x)) if pd.notna(x) else "Unknown"
            )
        else:
            # Create a default Data Source column if missing
            processed_df["Data Source"] = [f"Data_{i+1}" for i in range(len(processed_df))]

        # Filter out rows where all numeric columns are empty (0)
        # Keep rows that have at least one non-zero value in numeric columns
        numeric_columns = [col for col in processed_df.columns if col != "Data Source"]
        if numeric_columns:
            # Create a mask for rows that have at least one non-zero value in numeric columns
            has_data_mask = (processed_df[numeric_columns] != 0).any(axis=1)
            original_count = len(processed_df)
            processed_df = processed_df[has_data_mask].reset_index(drop=True)
            
            print(f"Filtered out {original_count - len(processed_df)} empty rows")
            
            # If all rows were empty, keep at least one row with default values
            if len(processed_df) == 0:
                print("All rows were empty, keeping one default row")
                processed_df = pd.DataFrame({
                    "Data Source": ["LF_1", "LH_1", "RF_1", "RH_1"],
                    "Maximum force [%BW]": [0, 0, 0, 0],
                    "Force-time integral [%BW*s]": [0, 0, 0, 0],
                    "Contact time/TO [ms]": [0, 0, 0, 0]
                })

        return processed_df
        
    except Exception as e:
        print(f"Error processing Excel data: {e}")
        # Return a minimal DataFrame with default values
        return pd.DataFrame({
            "Data Source": ["LF_1", "LH_1", "RF_1", "RH_1"],
            "Maximum force [%BW]": [0, 0, 0, 0],
            "Force-time integral [%BW*s]": [0, 0, 0, 0],
            "Contact time/TO [ms]": [0, 0, 0, 0]
        })

def add_additional_columns_to_sheet2(ws2, num_data_rows):
    """Add additional columns to the right of Sheet2."""
    # We have 4 original columns (A, B, C, D), so new columns start at E, F, G
    
    # Add empty column E with arrow in center
    if num_data_rows > 0:
        arrow_row_idx = (num_data_rows // 2) + 2 # +2 because Excel is 1-indexed
        arrow_cell = ws2.cell(row=arrow_row_idx, column=5, value="→")  # Column E
        arrow_cell.font = Font(size=14)
        arrow_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Also add a header for column E to make it clear
        arrow_header = ws2.cell(row=1, column=5, value="→")
        arrow_header.font = Font(size=12)
        arrow_header.alignment = Alignment(horizontal="center", vertical="center")

    # Add two more columns with headings
    # Column F: Weight bearing [%]
    weight_header = ws2.cell(row=1, column=6, value="Weight bearing [%]")
    weight_header.font = Font(bold=True)
    weight_header.alignment = Alignment(horizontal="center", vertical="center")
    
    # Column G: Asymmetry Index
    asym_header = ws2.cell(row=1, column=7, value="Asymmetery Index (L to R: SI)")
    asym_header.font = Font(bold=True)
    asym_header.alignment = Alignment(horizontal="center", vertical="center")
    
    # Center-align all data cells in the new columns
    for row_idx in range(2, num_data_rows + 2): # Start from row 2 for data
        # Column E (arrow column) - leave empty except for arrow
        if row_idx != arrow_row_idx:
            ws2.cell(row=row_idx, column=5, value="").alignment = Alignment(horizontal="center", vertical="center")
        # Column F (weight bearing)
        ws2.cell(row=row_idx, column=6, value="").alignment = Alignment(horizontal="center", vertical="center")
        # Column G (asymmetry index)
        ws2.cell(row=row_idx, column=7, value="").alignment = Alignment(horizontal="center", vertical="center")

def write_weight_bearing_formulae(ws2, num_data_rows):
    """Write weight bearing formulae for all rows with error handling."""
    try:
        # Weight bearing formula: IFERROR(ROUND((current_cell/SUM(range))*100, 0), "")
        # We need to group rows by 4 (LF, LH, RF, RH)
        
        for group_start in range(2, num_data_rows + 2, 4):  # Start from row 2, increment by 4
            if group_start + 3 <= num_data_rows + 1:  # Make sure we have 4 rows
                # Get the range for this group (4 rows)
                range_start = f"B{group_start}"
                range_end = f"B{group_start + 3}"
                range_reference = f"{range_start}:{range_end}"
                
                # Write formula for each row in the group
                for row_offset in range(4):
                    current_row = group_start + row_offset
                    if current_row <= num_data_rows + 1:
                        current_cell = f"B{current_row}"
                        # Enhanced formula with better error handling
                        weight_bearing_formula = f'=IFERROR(IF(SUM({range_reference})=0, 0, ROUND(({current_cell}/SUM({range_reference}))*100, 0)), "")'
                        cell = ws2.cell(row=current_row, column=6, value=weight_bearing_formula)  # Column F
                        cell.data_type = 'f'  # Explicitly set as formula
    except Exception as e:
        print(f"Error writing weight bearing formulae: {e}")
        # Fill with default values if formula writing fails
        for row in range(2, num_data_rows + 2):
            ws2.cell(row=row, column=6, value="0")

def write_asymmetry_formulae(ws2, num_data_rows):
    """Write asymmetry index formulae for LF and LH rows only with error handling."""
    try:
        # Asymmetry Index formula: IFERROR(ABS((x1-x3))/(AVERAGE(x1,x3)), "")
        # Only for LF and LH rows (first and second row in each group)
        
        # Calculate how many complete groups we have
        complete_groups = num_data_rows // 4
        remaining_rows = num_data_rows % 4
        
        # Write formulas for complete groups
        for group in range(complete_groups):
            group_start = 2 + (group * 4)
            
            # LF row (first row in group): abs(x1-x3)/mean(x1, x3) where x1=LF, x3=RF
            lf_row = group_start
            rf_row = group_start + 2
            if rf_row <= num_data_rows + 1:  # Make sure RF row exists
                # Enhanced formula with better error handling for division by zero
                lf_formula = f'=IFERROR(IF(AVERAGE(B{lf_row},B{rf_row})=0, 0, ABS((B{lf_row}-B{rf_row}))/AVERAGE(B{lf_row},B{rf_row})), "")'
                cell = ws2.cell(row=lf_row, column=7, value=lf_formula)  # Column G
                cell.data_type = 'f'  # Explicitly set as formula
            
            # LH row (second row in group): abs(x2-x4)/mean(x2, x4) where x2=LH, x4=RH
            lh_row = group_start + 1
            rh_row = group_start + 3
            if rh_row <= num_data_rows + 1:  # Make sure RH row exists
                # Enhanced formula with better error handling for division by zero
                lh_formula = f'=IFERROR(IF(AVERAGE(B{lh_row},B{rh_row})=0, 0, ABS((B{lh_row}-B{rh_row}))/AVERAGE(B{lh_row},B{rh_row})), "")'
                cell = ws2.cell(row=lh_row, column=7, value=lh_formula)  # Column G
                cell.data_type = 'f'  # Explicitly set as formula
        
        # Handle remaining rows if we don't have complete groups
        if remaining_rows > 0:
            group_start = 2 + (complete_groups * 4)
            
            # If we have at least 2 rows, we can calculate LF asymmetry
            if remaining_rows >= 2:
                lf_row = group_start
                rf_row = group_start + 2
                if rf_row <= num_data_rows + 1:
                    lf_formula = f'=IFERROR(IF(AVERAGE(B{lf_row},B{rf_row})=0, 0, ABS((B{lf_row}-B{rf_row}))/AVERAGE(B{lf_row},B{rf_row})), "")'
                    cell = ws2.cell(row=lf_row, column=7, value=lf_formula)  # Column G
                    cell.data_type = 'f'  # Explicitly set as formula
            
            # If we have at least 3 rows, we can calculate LH asymmetry
            if remaining_rows >= 3:
                lh_row = group_start + 1
                rh_row = group_start + 3
                if rh_row <= num_data_rows + 1:
                    lh_formula = f'=IFERROR(IF(AVERAGE(B{lh_row},B{rh_row})=0, 0, ABS((B{lh_row}-B{rh_row}))/AVERAGE(B{lh_row},B{rh_row})), "")'
                    cell = ws2.cell(row=lh_row, column=7, value=lh_formula)  # Column G
                    cell.data_type = 'f'  # Explicitly set as formula
                    
    except Exception as e:
        print(f"Error writing asymmetry formulae: {e}")
        # Fill with default values if formula writing fails
        for row in range(2, num_data_rows + 2):
            ws2.cell(row=row, column=7, value="0")

def apply_coloring(ws2, num_data_rows):
    """Apply coloring to Data Source (column A) and Weight bearing columns (column F)."""
    # Color cache for consistent coloring
    color_cache = {}
    
    # Apply coloring to Data Source (column A) and Weight bearing columns (column F)
    for row_idx in range(2, num_data_rows + 2): # Start from row 2 for data
        # Determine group number based on row index - group by 4 rows (LF, LH, RF, RH)
        group_num = ((row_idx - 2) // 4) + 1  # Group 1: rows 2-5, Group 2: rows 6-9, etc.
        
        if group_num not in color_cache:
            color_cache[group_num] = {
                'bright': get_color_for_number(group_num),
                'dim': get_dim_color_for_number(group_num)
            }
        
        # Apply bright color to Data Source column (column A)
        data_source_cell = ws2.cell(row=row_idx, column=1)
        data_source_cell.fill = PatternFill(start_color=color_cache[group_num]['bright'], 
                                           end_color=color_cache[group_num]['bright'], 
                                           fill_type='solid')
        
        # Apply dim color to Weight bearing column (column F)
        weight_bearing_cell = ws2.cell(row=row_idx, column=6)
        weight_bearing_cell.fill = PatternFill(start_color=color_cache[group_num]['dim'], 
                                              end_color=color_cache[group_num]['dim'], 
                                              fill_type='solid')

def add_summary_averages_table(ws2, num_data_rows):
    """Add a summary table with averages for LF, LH, RF, RH groups below the main data."""
    # Add 3-4 rows gap after the main data
    gap_start_row = num_data_rows + 4  # Main data + gap
    
    # Add heading "Summary"
    heading_cell = ws2.cell(row=gap_start_row, column=1, value="Summary")
    heading_cell.font = Font(bold=True, size=14)
    heading_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create summary table
    table_start_row = gap_start_row + 2
    
    # Define column headers for the summary table
    summary_headers = ["", "Maximum force [%BW]", "Force-time integral [%BW*s]", "Contact time/TO [ms]", "Weight bearing [%]"]
    
    # Write headers
    for col_idx, header in enumerate(summary_headers):
        header_cell = ws2.cell(row=table_start_row, column=col_idx + 1, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows for each group (LF, LH, RF, RH)
    row_idx = table_start_row + 1
    for group_idx, prefix in enumerate(['LF', 'LH', 'RF', 'RH']):
        # Group name - make it bold and center-aligned
        group_cell = ws2.cell(row=row_idx, column=1, value=prefix)
        group_cell.font = Font(bold=True)
        group_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Calculate the starting row for this group in the data
        # Group 0 (LF): rows 2, 6, 10, 14, 18... (every 4th row starting from 2)
        # Group 1 (LH): rows 3, 7, 11, 15, 19... (every 4th row starting from 3)
        # Group 2 (RF): rows 4, 8, 12, 16, 20... (every 4th row starting from 4)
        # Group 3 (RH): rows 5, 9, 13, 17, 21... (every 4th row starting from 5)
        group_start_row = 2 + group_idx  # Starting row for this group
        
        # Create formulas for each metric - only for rows belonging to this group
        # Maximum force average formula (Column B)
        formula_range = []
        for data_row in range(group_start_row, num_data_rows + 2, 4):  # Every 4th row starting from group_start_row
            if data_row <= num_data_rows + 1:  # Make sure we don't exceed data rows
                col_letter = "B"  # Column B for Maximum force
                cell_ref = f"{col_letter}{data_row}"
                formula_range.append(cell_ref)
        
        if formula_range:
            avg_formula = f'=ROUNDDOWN(AVERAGE({",".join(formula_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(formula_range)}),2)'
            formula_cell = ws2.cell(row=row_idx, column=2, value=avg_formula)
            formula_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Force-time integral average formula (Column C)
        formula_range = []
        for data_row in range(group_start_row, num_data_rows + 2, 4):  # Every 4th row starting from group_start_row
            if data_row <= num_data_rows + 1:  # Make sure we don't exceed data rows
                col_letter = "C"  # Column C for Force-time integral
                cell_ref = f"{col_letter}{data_row}"
                formula_range.append(cell_ref)
        
        if formula_range:
            avg_formula = f'=ROUNDDOWN(AVERAGE({",".join(formula_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(formula_range)}),2)'
            formula_cell = ws2.cell(row=row_idx, column=3, value=avg_formula)
            formula_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Contact time average formula (Column D)
        formula_range = []
        for data_row in range(group_start_row, num_data_rows + 2, 4):  # Every 4th row starting from group_start_row
            if data_row <= num_data_rows + 1:  # Make sure we don't exceed data rows
                col_letter = "D"  # Column D for Contact time
                cell_ref = f"{col_letter}{data_row}"
                formula_range.append(cell_ref)
        
        if formula_range:
            avg_formula = f'=ROUNDDOWN(AVERAGE({",".join(formula_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(formula_range)}),2)'
            formula_cell = ws2.cell(row=row_idx, column=4, value=avg_formula)
            formula_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Weight bearing average formula (Column F)
        formula_range = []
        for data_row in range(group_start_row, num_data_rows + 2, 4):  # Every 4th row starting from group_start_row
            if data_row <= num_data_rows + 1:  # Make sure we don't exceed data rows
                col_letter = "F"  # Column F for Weight bearing
                cell_ref = f"{col_letter}{data_row}"
                formula_range.append(cell_ref)
        
        if formula_range:
            avg_formula = f'=ROUNDDOWN(AVERAGE({",".join(formula_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(formula_range)}),2)'
            formula_cell = ws2.cell(row=row_idx, column=5, value=avg_formula)
            formula_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_idx += 1
    
    # Auto-adjust column widths for the summary table
    for col in range(5):
        header_value = summary_headers[col]
        if header_value:
            header_length = len(str(header_value)) + 2
            adjusted_width = min(max(header_length, 10), 50)
            ws2.column_dimensions[get_column_letter(col + 1)].width = adjusted_width

def add_forelimb_hindlimb_summary(ws2, num_data_rows):
    """Add a summary table for forelimb/hindlimb asymmetry index averages below the main calculations table."""
    # Add 1-2 rows gap after the main calculations table
    gap_start_row = num_data_rows + 4 + 2 + 4  # Main data + gap + summary table + gap
    
    # Create simple 2-column table (no heading row)
    table_start_row = gap_start_row + 1
    
    # Write headers
    header_cell1 = ws2.cell(row=table_start_row, column=1, value="")
    header_cell1.font = Font(bold=True)
    header_cell1.alignment = Alignment(horizontal="center", vertical="center")
    
    header_cell2 = ws2.cell(row=table_start_row, column=2, value="Asym Index(L to R: SI)")
    header_cell2.font = Font(bold=True)
    header_cell2.alignment = Alignment(horizontal="center", vertical="center")
    
    # Forelimb row (LF + LH asymmetry averages)
    forelimb_row = table_start_row + 1
    forelimb_cell = ws2.cell(row=forelimb_row, column=1, value="Forelimb")
    forelimb_cell.font = Font(bold=True)
    forelimb_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create dynamic formula for forelimb - only LF rows have asymmetry index values
    forelimb_range = []
    for row in range(2, num_data_rows + 2, 4):  # Every 4th row starting from 2 (LF rows)
        if row <= num_data_rows + 1:
            forelimb_range.append(f"G{row}")
    
    if forelimb_range:
        forelimb_formula = f'=ROUNDDOWN(AVERAGE({",".join(forelimb_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(forelimb_range)}),2)'
    else:
        forelimb_formula = '""'
    
    forelimb_cell = ws2.cell(row=forelimb_row, column=2, value=forelimb_formula)
    forelimb_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Hindlimb row (RF + RH asymmetry averages)
    hindlimb_row = table_start_row + 2
    hindlimb_cell = ws2.cell(row=hindlimb_row, column=1, value="Hindlimb")
    hindlimb_cell.font = Font(bold=True)
    hindlimb_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create dynamic formula for hindlimb - only LH rows have asymmetry index values
    hindlimb_range = []
    for row in range(3, num_data_rows + 2, 4):  # Every 4th row starting from 3 (LH rows)
        if row <= num_data_rows + 1:
            hindlimb_range.append(f"G{row}")
    
    if hindlimb_range:
        hindlimb_formula = f'=ROUNDDOWN(AVERAGE({",".join(hindlimb_range)}),2)&"±"&ROUNDDOWN(STDEV({",".join(hindlimb_range)}),2)'
    else:
        hindlimb_formula = '""'
    
    hindlimb_cell = ws2.cell(row=hindlimb_row, column=2, value=hindlimb_formula)
    hindlimb_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 30

def set_column_widths(ws2, df):
    """Set column widths based on the content of the DataFrame."""
    # Set column widths for the main data columns
    for col_idx, col_name in enumerate(df.columns):
        # Calculate width based on header length and max data length
        header_length = len(str(col_name))
        max_data_length = df[col_name].astype(str).str.len().max()
        width = max(header_length, max_data_length) + 2  # Add padding
        width = min(max(width, 10), 50)  # Min 10, Max 50 characters
        ws2.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    # Set widths for additional columns
    ws2.column_dimensions['E'].width = 15 # Arrow column
    ws2.column_dimensions['F'].width = 20 # Weight bearing column
    ws2.column_dimensions['G'].width = 35 # Asymmetry index column

def get_color_for_number(n):
    """Generate a unique pastel-like color for each number using HSL."""
    hue = (n * 137) % 360 / 360.0  # golden angle for good distribution
    lightness = 0.75
    saturation = 0.9
    r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
    return f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"

def get_dim_color_for_number(n):
    """Generate a dimmer version of the color for each number using HSL."""
    hue = (n * 137) % 360 / 360.0  # golden angle for good distribution
    lightness = 0.75  # Much lighter (dimmer)
    saturation = 0.5  # Less saturated
    r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
    return f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"


def calculate_age_years_months(date_of_birth_str, date_of_visit_str=None):
    """
    Calculate age in years and months from date of birth.
    
    Args:
        date_of_birth_str: Date of birth in format like "4/1/18" or "4/1/2018"
        date_of_visit_str: Date of visit in format like "8/21/25" or "8/21/2025" (optional)
    
    Returns:
        Tuple of (years, months)
    """
    try:
        # Parse date of birth
        dob_parts = date_of_birth_str.split('/')
        month, day, year = dob_parts[0], dob_parts[1], dob_parts[2]
        
        # Handle 2-digit years (assume 20xx for years < 50, 19xx for years >= 50)
        if len(year) == 2:
            year_int = int(year)
            if year_int < 50:
                year = f"20{year}"
            else:
                year = f"19{year}"
        
        dob = date(int(year), int(month), int(day))
        
        # Use date of visit if provided, otherwise use current date
        if date_of_visit_str:
            visit_parts = date_of_visit_str.split('/')
            visit_month, visit_day, visit_year = visit_parts[0], visit_parts[1], visit_parts[2]
            
            # Handle 2-digit years for visit date
            if len(visit_year) == 2:
                visit_year_int = int(visit_year)
                if visit_year_int < 50:
                    visit_year = f"20{visit_year}"
                else:
                    visit_year = f"19{visit_year}"
            
            visit_date = date(int(visit_year), int(visit_month), int(visit_day))
        else:
            visit_date = date.today()
        
        # Calculate years and months
        years = visit_date.year - dob.year
        months = visit_date.month - dob.month
        
        # Adjust if birthday hasn't occurred yet this year
        if visit_date.day < dob.day:
            months -= 1
        
        # Adjust years if months is negative
        if months < 0:
            years -= 1
            months += 12
            
        return years, months
        
    except Exception as e:
        print(f"Error calculating age: {e}")
        return None, None


